import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# Открываем книгу Excel
workbook = openpyxl.load_workbook('rts_test.xlsx')

# Получаем активный лист
sheet = workbook.active

# Создаем новый лист
new_sheet = workbook.create_sheet("sort")

# Создаем заголовки для нового листа
#new_sheet.append(["Менеджер", "Документ", "Сумма документа"])

# Считываем данные из Excel в DataFrame
data = sheet.values
df = pd.DataFrame(data)

# Сортируем данные по менеджерам и добавляем их на новый лист
for manager, group in df.groupby(df.iloc[:, 5]):
    for row in dataframe_to_rows(group, index=False, header=False):
        new_sheet.append([manager, row[0], row[9]])

# Сохраняем изменения
workbook.save('rts_test.xlsx')